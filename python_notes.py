from sys import argv  # variable arguments
###################################################
script, first, second, third = argv
print('{first} & {second} & {third}')
# When called: python hello.py Wayne Fucking Rooney
from os.path import exists  EXISTS
print(f"Does the output file exists? {exists(out_file)}")
things = ['United', 'Zaha', 'Wazza', 'Chalton', 'Pogba', 'Trafford']  JOIN
print(' '.join(things))  # United Zaha Wazza Chalton Pogba Trafford
print('#'.join(things[1:4]))  # Zaha#Wazza#Chalton
teams = {'UTD': 'United', 'CHE': 'Chelsea', 'ARS': 'Arsenal', 'LIV': 'Liverpool'} DICT.GET
team = teams.get('BAR')
print(team)  # None
team = teams.get('BAR', 'Does Not Exist')
print(team)  # Does Not Exist
numbers = [1, 2, 3] LIST METHODS
numbers.append(4)  # [1, 2, 3, 4]
numbers.insert(1, 77)  # [1, 77, 2, 3]
numbers.remove(2)  # [1, 77, 3, 4]
numbers.clear()  # []
numbers.pop()  # [1, 77, 3]
print(numbers.index(77))  # 1
print(numbers.count(77))  # 1
numbers.sort()  # [1, 3 ,77]
numbers.reverse()  # [77, 3 ,1]
numbers2 = numbers.copy()
# Inheritance:  INHERITANCE


class Patruped():
  def mers(self):
    print("merge in 4 membre")


class Caine(Patruped):
  def __init__(self, name):
    self.name = name

  def latra(self):
    print("{} face ham ham".format(self.name))


class Pisica(Patruped):
  def __init__(self, name):
    self.name = name

  def miaun(self):
    print("{} face miaun miaun".format(self.name))


oscar = Caine(name='Oscar')
oscar.latra()  # Oscar face ham ham
codita = Pisica(name='Codita')
codita.miaun()  # Codita face miaun miaun
import pathlib  PATHLIB
# sau
from pathlib import Path

path = Path("Jucatori")
path.mkdir()  # creeaza folderul jucatori
path.rmdir()  # Sterge folderul jucatori
from pathlib import Path  PATHLIB

path = Path()
for file in path.glob('*.*'):
  path.glob('*')
  path.glob('*.xls')
  print(file)
import openpyxl as xl  # shortcut OPENPYXL
from openpyxl.chart import BarChart, Reference  # for charts


def process_workbook(filename):
  wb = xl.load_workbook(filename)  # opening an excel file
  sheet = wb['Sheet1']  # opening a sheet

  for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)  # cell as an object
    corrected_price = cell.value * 2  # cell.value takes the value from cell
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

  values = Reference(sheet,  # taking the values
                     min_row=2,
                     max_row=sheet.max_row,
                     min_col=4,
                     max_col=4)

  chart = BarChart()
  chart.add_data(values)  # adding the values
  sheet.add_chart(chart, 'e2')  # adding the chart
  wb.save(filename)


process_workbook('transactions.xlsx')


class Parent(object):
  INHERITANCE

  def override(self):
    print("PARENT override()")

  def implicit(self):
    print("PARENT implicit()")

  def altered(self):
    print("PARENT altered()")


class Child(Parent):
  def override(self):
    print("CHILD override()")

  def altered(self):
    print("CHILD, BEFORE PARENT altered()")
    super(Child, self).altered()
    print("CHILD, AFTER PARENT altered()")


dad = Parent()
son = Child()
dad.implicit()  # PARENT implicit()
son.implicit()  # PARENT implicit()
dad.override()  # PARENT override()
son.override()  # CHILD override()
dad.altered()  # PARENT altered()
son.altered()   # CHILD, BEFORE PARENT altered()
# PARENT altered()
# CHILD, AFTER PARENT altered()
print(int(28.71))  # 28 ROUND TRUNCATE
print(round(93.34836, 2))  # 93.35
print(round(19.47, 0))  # 19.0
print(int(19.47))  # 18
print(round(28793.54836, 0))  # 28794.0
print(round(28793.54836, 1))  # 28793.5
print(round(28793.54836, -1))  # 28790.0
print() == print(endl="\n") END =
print(12, 33, False, sep=">")  # 12>33>False  SEP=
print('{0} {1} {0}'.format(True, False))  # True False True FORMAT
print('{0:04} {1:.5f}'.format(5, 5))  # 0005 5.00000  FORMAT
# *                                #  FORMAT
print('{0:>4} {1:>32}'.format('*', '#'))
for i in range(16):
  FORMATED OUTPUT
  print('{0:3} {1:16}'.format(i, 10**i))
size = int(input("Please enter the table size: "))  FORMATED TABLE
# Print a size x size multiplication table
for row in range(1, size + 1):
  for column in range(1, size + 1):
    product = row * column  # Compute product
    print('{0:4}'.format(product), end='')  # Display product
  print()
# I will inject text here and 34 here FORMAT
print("I will inject %s here and %d here" % ('text', 34))
o_lista = ['ala', 'bala', 'portocala']  ENUMERATE
for elem in enumerate(o_lista, start=1):
  print(elem, end="")  # (1, 'ala')(2, 'bala')(3, 'portocala')
lista1 = [1, 2, 3, 4, 5, 6] ZIP
lista2 = ['a', 'b', 'c', 'd', 'e']
lista3 = [100, 200, 300, 400]

for item in zip(lista1, lista2, lista3):
  print(item)  # (1, 'a', 100)
#                   (2, 'b', 200)
#                   (3, 'c', 300)
#                   (4, 'd', 400)
# sau new_list(zip(lista1, lista2, lista3))


def cauta_caine(x, y):  RETURN
  if x.lower() in y.lower():
    return True
  else:
    return False
# sau


def cauta_cainev2(cuvant, sir):
  return cuvant.lower() in sir.lower()


def myfunc(**kwargs): ARGS KWARGS
  print(kwargs)  # (40, 60, 100, 1, 43)
  if 'utd' in kwargs:
    print(f"Echipa mea preferata este {kwargs['utd']}.")
  elif 'mci' in kwargs or 'liv' in kwargs:
    print(f"Echipele rivale sunt {kwargs['mci']},{kwargs['liv']}.")


# Echipa mea preferata este United.
myfunc(utd='United', mci='City', liv='Liverpool')


def myfunc(*args, **kwargs):
  print(f"I would like {args[1]} {kwargs['car']}.")


# I would like 20 Mercedeses.
myfunc(10, 20, 30, food='Pizzas', car='Mercedeses')


def printare(nume): MAP FILTER
  return f'{nume} e BO$$'


def patrat(numar):
  return numar ** 2


def check_nume(numele):
  return 'elena' in numele.lower()


nume_fete = ['Elena Popescu', 'Ivan Elena', 'Simona Urs', 'Raluca Elena']
numere = [1, 2, 3, 4, 5]
catalog = ['Costel', 'Nicu', 'Gigel']

lista = list(map(printare, catalog))
print(lista)  # ['Costel e BO$$', 'Nicu e BO$$', 'Gigel e BO$$']

for item in map(patrat, numere):
  print(item, end=" ")  # 1 4 9 16 25

lista3 = list(filter(check_nume, nume_fete))
print('\n', lista3)  # ['Elena Popescu', 'Ivan Elena', 'Raluca Elena']
numere = [2, 5, 7, 8, 4]  LAMBDA
lista = list(map(lambda num: num ** 2, numere))
print(lista)  # [4, 25, 49, 64, 16]

titulaturi = ['Raluca', 'Elena', 'Bianca', 'Casandra']
lista2 = list(filter(lambda nume: 'ca' in nume.lower(), titulaturi))
print(lista2)  # ['Raluca', 'Bianca', 'Casandra']
name = 'Henrik' CLASS


class Caine:

  # Class object atributes: Same for all the instances (objects)
  species = 'mammal'

  def __init__(self, breed, name):
    self.breed = breed
    self.name = name
    self.specius = Caine.species + 'ius'

  def ce_face(self, numar):
    print(f'{self.name} face ham ham de {numar} ori')


caini = {'Igor': 'pitbull', 'Lord': 'lup', 'Oscar': 'husky'}

for nume, rasa in caini.items():
  nume = Caine(rasa, nume)
  nume.ce_face(4)
  print(nume.specius)
  print(f'Si este rasa: {nume.breed}, din regnul {nume.species}')
# Polymorphism  POLYMORPHISM


class Dog():

  def __init__(self, name):
    self.name = name

  def speak(self):
    return self.name + ' says wof'


class Cat():

  def __init__(self, name):
    self.name = name

  def speak(self):
    return self.name + ' says meow'


niko = Dog('niko')
felix = Cat('felix')
print(niko.speak())
print(felix.speak())

for pet in [niko, felix]:
  print(type(pet))
  print(pet.speak())


def pet_speak(pet):
  print(pet.speak())


print(pet_speak(niko))
print(pet_speak(felix))


class Book:
  __STR__ __LEN__ __DEL__

  def __init__(self, titlu, autor, pagini):
    self.titlu = titlu
    self.autor = autor
    self.pagini = pagini

  def __str__(self):
    return f'{self.titlu}, scrisa de {self.autor}'

  def __len__(self):
    return self.pagini

  def __del__(self):
    print('O carte a fost stearsa')


carte = Book("Cool", 'BRATU', 100)
print(carte)
print(len(carte))
del carte


def funct():  __NAME__ __MAIN__
  print('FUNC() IN ONE.PY')


print('TOP LEVEL IN ONE.PY')

if __name__ == '__main__':
  print('ONE.PY is being run directly')
else:
  print('ONE.PY has been imported')

# two.py
import one

print('TOP LEVEL IN TWO.PY')

one.funct()

if __name__ == '__main__':
  print('TWO.PY is run directly')
else:
  print('Two.py has been imported')
# def cap_text(text): UNITTEST
#     return text.title()

import unittest
import cap


class TestCap(unittest.TestCase):

  def test_one_word(self):
    text = 'python'
    result = cap.cap_text(text)
    self.assertEqual(result, 'Python')

  def test_multiple_words(self):
    text = 'molty python'
    result = cap.cap_text(text)
    self.assertEqual(result, 'Molty Python')


if __name__ == '__main__':
  unittest.main()
# def cap_adu(*args): UNITTEST
#    return sum(args)

import unittest
import cap


class TestAdunare(unittest.TestCase):

  def test_two_nums(self):
    a = 4
    b = 6
    result = cap.cap_adu(a, b)
    self.assertEqual(result, 10)

  def test_three_nums(self):
    a = 5
    b = 60
    c = 4
    result = cap.cap_adu(a, b, c)
    self.assertEqual(result, 69)


if __name__ == '__main__':
  unittest.main()


def hello(name='Dragos'): FUNCSAPTION
  print('The hello() func has been executed')

  def greet():
    return '\t This is a greet funct inside hello!'

  def welcome():
    return '\t this is inside hello'

  print('I am going to return a funct')
  if name == 'Dragos':
    return greet
  else:
    return welcome


my_new_funct = hello('Dragos')
print(my_new_funct())


def hello():  FUNCSAPTION
  return 'Hi Dragos!'


def other(some_def_func):
  print('Other code runs here!')
  print(some_def_func())


other(hello)


def new_decorator(original_func): DECORATORS

  def wrap_func():
    print('some extra code, before the original function')
    original_func()
    print('Some extra code, after the original func')
  return wrap_func


def func_needs_decorator():
  print('I want to be decorated!')


@new_decorator
def func_needs_decorator():
  print('I want to be decorated!')


func_needs_decorator()


def inmultire_impartire(original_func): DECORATORS

  def wrap_func(a, b):
    print(f'produsul a 2 numere este{a * b}')
    original_func(a, b)
    print(f'catul a 2 numere este{a / b}')
  return wrap_func


def operatii_matematice(a, b):
  print(f'suma a 2 numere este{a + b}')


@inmultire_impartire
def operatii_matematice(a, b):
  print(f'suma a 2 numere este {a + b}')


operatii_matematice(3, 5)


def create_cubes(n):  GENERATORS
  for x in range(n):
    yield x**3


generatorul = create_cubes(9)
copie_gen = create_cubes(9)
lista = list(generatorul)
print(type(generatorul))  # <class 'generator'>
print(lista)  # [0, 1, 8, 27, 64, 125, 216, 343, 512]
print(next(copie_gen))  # 0
print(next(copie_gen))  # 1
print(next(copie_gen))  # 8
s = 'Hello' ITER
s_iter = iter(s)
print(next(s_iter))  # H
print(next(s_iter))  # e
from collections import Counter COUNTER

lista = [1, 2, 3, 1, 2, 2, 2, 3, 1, 2, 8]
sir = 'asdadfafasdasdafsdasfdvn'
text = 'Ana are mere si pere si mere da mere nu pere poate mere'
cuvinte = text.split()
print(Counter(lista))  # Counter({2: 5, 1: 3, 3: 2, 8: 1})
# Counter({'a': 7, 'd': 6, 's': 5, 'f': 4, 'v': 1, 'n': 1})
print(Counter(sir))
print(Counter(cuvinte))  # Counter({'mere': 4, 'si': 2, 'pere': 2, 'Ana': 1
# , 'are': 1, 'da': 1, 'nu': 1, 'poate': 1})
# c.clear() - Resets all counts COUNTER
# list(c) - list unique elements
# set(c) - convert to  set
# dict(c) - convert tot a regular dictionary
# c.items() - comvert to a list of (elem, cnt) pairts
# Counter(dict(list_of_pairs)) - convert from a list of (elem, cnt) pairs
# c.most_common() [:-n-1:-1] - n Least common elements
# c += Counter() - Remove zero and negative counts
from collections import defaultdict DEFAULT DICT

d = defaultdict(lambda: 0)
d['one']
d['two'] = 2
# defaultdict(<function <lambda> at 0x0000000002201828>, {'a': 0, 'b': 2})
print(d)
from collections import OrderedDict ORDERED DICT

dictionar_normal1 = {}
dictionar_normal2 = {}
dictionar_normal1['a'] = '1'
dictionar_normal1['b'] = '2'
dictionar_normal2['b'] = '2'
dictionar_normal2['a'] = '1'
print(dictionar_normal1 == dictionar_normal2)  # True
dictionar_ordonat1 = OrderedDict()
dictionar_ordonat2 = OrderedDict()
dictionar_ordonat1['a'] = '1'
dictionar_ordonat1['b'] = '2'
dictionar_ordonat2['b'] = '2'
dictionar_ordonat2['a'] = '1'
print(dictionar_ordonat1 == dictionar_ordonat2)  # False
from collections import namedtuple  NAMEDTUPLE

standard_tuple = (1, 2, 3)
print(standard_tuple[1])  # 2

Dog = namedtuple('Dog', 'varsta rasa nume')
sam = Dog(varsta=14, rasa='Lab', nume='Sammy')
print(sam.nume)  # Sammy
print(sam[2])  # Sammy
print(sam.varsta)  # 14
import datetime DATETIME
t = datetime.time(5, 25, 1)
print(t)  # 05:25:01
print(t.hour)  # 5
print(datetime.time.min)  # 00:00:00
print(datetime.time.max)  # 23:59:59.999999
import datetime DATETIME

today = datetime.date.today()
print(today)  # 2019-10-14
print(today.year)  # 2019
print(datetime.date.min)  # 0001-01-01
print(datetime.date.max)  # 9999-12-31
d1 = datetime.date(2015, 3, 11)
print(d1)  # 2015-03-11
d2 = d1.replace(year=1990)
print(d2)  # 1990-03-11
print(d1 - d2)  # 9131 days, 0:00:00
import pdb  PDB

x = [1, 3, 4]
y = 2
z = 3
result = y + z
print(result)
pdb.set_trace()
result2 = y + x
print(result2)
import timeit TIMEIT

print(timeit.timeit('"-".join(str(n) for n in range(100))', number=10000))
# 0.402033814
print(timeit.timeit('"-".join([str(n) for n in range(100)])', number=10000))
# 0.345903018
print(timeit.timeit('"-".join(map(str,range(100)))', number=10000))
# 0.222068923
import re FINDALL

text_phrase = 'sdsd..sssddd...sdddsddd...dsds...dsssss...sdddd'
text_patterns = ['sd*',  # s followed by zero or more d's
                 'sd+',  # s followed by one or more d's
                 'sd?',  # s followed by zero or one d's
                 'sd{3}',  # s followed by three d's
                 'sd{2,3}',  # s followed by two to three d's
                 ]


def multi_re_find(patterns, phrase):
  for pattern in patterns:
    print(f'searching the phrase using the re check: {pattern}')
    print(re.findall(pattern, phrase))
    print('\n')


multi_re_find(text_patterns, text_phrase)
import re FINDALL

test_phrase = 'This is a string! But it has punctuation. How can we remove it?'
x = re.findall('[^!.? ]+', test_phrase)
print(x)
print(hex(13))  # 0xd NUMBERS
print(bin(14))  # 0b1110
print(pow(2, 3, 3))  # (2^3)%3 = 2
print(abs(-2))  # 2
print(round(3.1))  # 3.0 (Always float)
print(round(1.2345, 2))  # 1.23
s = 'United'  STRINGS
print(s.isalnum())  # True
print(s.islower())  # False
print(s.isalpha())  # True
s = set() SETS
s.add(1)
s.add(2)
print(s)  # {1, 2}
s.clear()
print(s)  # set()
s = {1, 2, 3}
sc = s.copy()
print(sc)  # {1, 2, 3}
s.add(4)
print(s)  # {1, 2, 3, 4}
print(s.difference(sc))  # {4}
s1 = {1, 2, 3}
s2 = {1, 4, 5}
s1.difference_update(s2)
print(s1)  # {2, 3}
s.discard(2)  # {1, 2, 3, 4}
print(s)  # {1, 3, 4}
s1 = {1, 2, 3}
s2 = {1, 2, 4}
print(s1.intersection(s2))  # {1, 2}
s1.intersection_update(s2)
print(s1)  # {1, 2}
s1 = {1, 2}
s2 = {1, 2, 4}
s3 = {5}
print(s1.isdisjoint(s2))  # False
print(s1.isdisjoint(s3))  # True
print(s1.issubset(s2))  # True
print(s2.issuperset(s1))  # True
print(s1.symmetric_difference(s2))  # {4}
print(s1.union(s2))  # {1, 2, 4}
s1 = {1, 2}
s2 = {1, 2, 4}
print(s1.update(s2))
z = {k: v**2 for k, v in zip(['a', 'b'], range(2))} DICTIONARIES
print(z)  # {'a': 0, 'b': 1}
print(ord('a'))  # 97 ORD
print(ord('b'))  # 98
print(ord('A'))  # 65
dividend = int(input('Enter dividend: ')) COMPREHENSIVE IF
divisor = int(input('Enter divisor: '))
msg = dividend / divisor if divisor != 0 else 'Error, cannot divide by zero'
print(msg)
n = int(input("Enter a number: "))  COMPREHENSIVE IF
print('|', n, '| = ', (-n if n < 0 else n), sep='')
nume = 'dragos' if 'e' in 'Daniel' else 'Bratu' COMPREHENSIVE IF
print(nume)
count = sum = 0 WHILE ELSE
print('Please provide five nonnegative numbers when prompted')
while count < 5:
  # Get value from the user
  val = float(input('Enter number: '))
  if val < 0:
    print('Negative numbers not acceptable! Terminating')
    break
  count += 1
  sum += val
else:
  print('Average =', sum / count)
import turtle TURTLE


def o_linie(numar, culoare):
  for i in range(numar):
    turtle.pencolor(culoare)
    turtle.forward(200)  # Start
    turtle.left(90)
    turtle.forward(1)
    turtle.left(90)
    turtle.forward(200)
    turtle.right(90)
    turtle.forward(1)
    turtle.right(90)


o_linie(5, 'red')
o_linie(5, 'blue')
o_linie(5, 'red')
o_linie(5, 'blue')
o_linie(5, 'red')

turtle.hideturtle()
turtle.exitonclick()
from time import perf_counter, sleep  TIME MODULE
print("Enter your name: ", end="")
start_time = perf_counter()
name = input()
elapsed = perf_counter() - start_time
print(name, "it took you", elapsed, "seconds to respond")
for count in range(10, -1, -1):  # Range 10, 9, 8, ..., 0
  print(count)  # Display the count
  sleep(1)  # Suspend execution for 1 second
import sys  SYS.EXIT
sum = 0
while True:
  x = int(input('Enter a number (999 ends):'))
  if x == 999:
    sys.exit(0)
  sum += x
print('Sum is', sum)
x1 = eval(input('Entry x1? '))  # 4 EVAL
print('x1 =', x1, ' type:', type(x1))  # x1 = 4  type: <class 'int'>
x2 = eval(input('Entry x2? '))  # 4.0
print('x2 =', x2, ' type:', type(x2))  # x2 = 4.0  type: <class 'float'>
x3 = eval(input('Entry x3? '))  # 'x1'
print('x3 =', x3, ' type:', type(x3))  # x3 = x1  type: <class 'str'>
x4 = eval(input('Entry x4? '))  # x1
print('x4 =', x4, ' type:', type(x4))  # x4 = 4  type: <class 'int'>
x5 = eval(input('Entry x5? '))  # x6
print('x5 =', x5, ' type:', type(x5))  # NameError: name 'x6' is not defined
x = 4e3 POWER OF 10
print(x)  # 4000.0"


def gcd(m, n):  CMMDC
  min = m if m < n else n
  largest_factor = 1
  for i in range(1, min + 1):
    if m % i == 0 and n % i == 0:
      largest_factor = i  # Found larger factor
  return largest_factor


def get_int():
  return int(input("Please enter an integer: "))


def main():
  n1 = get_int()
  n2 = get_int()
  print("gcd(", n1, ",", n2, ") = ", gcd(n1, n2), sep="")


main()
from math import sqrt PRIME NUMBERS


def is_prime(n):
  root = round(sqrt(n)) + 1
  # Try all potential factors from 2 to the square root of n
  for trial_factor in range(2, root):
    if n % trial_factor == 0:  # Is it a factor?
      return False  # Found a factor
  return True  # No factors found


def main():
  max_value = int(input("Display primes up to what value? "))
  for value in range(2, max_value + 1):
    if is_prime(value):  # See if value is prime
      print(value, end=" ")  # Display the prime number


main()
word = "ABCD" RJUST / LJUST
print(word.rjust(10, "*"))  # ******ABCD
print(word.rjust(3, "*"))  # ABCD
print(word.rjust(15, ">"))  # >>>>>>>>>>>ABCD
print(word.rjust(10))  # ABCD
sir = "21ABCD1A23"  STRIP()
print(sir.strip('123'))  # ABCD1A
len(sir) == sir.__len__()
sir[2] == sir.__getitem__(2)
pairs = [(x, y) for x in [1, 2, 3] for y in ['a', 'b']] ZIP
print(pairs)  # [(1, 'a'), (1, 'b'), (2, 'a'), (2, 'b'), (3, 'a'), (3, 'b')]
print([p for p in range(2, 80) if not [x for x in range(2, p) if p % x == 0]])  PRIME NUMBERS
lst = [2 * i for i in range(6)] PRINTING
print(lst)  # Typical list printing
print(*lst)  # Print just the list elements
# Print the list in a special way
print(*lst, sep=" and ", end="--that's all folks!\n")
lista = ['a', 'b', 'c'] DICTIONARIES
numere = [1, 2, 3]
dictionar = dict(zip(lista, numere))
print(dictionar)
a = {1, 2, 3, 4}  SETS OPERATIONS
b = {3, 4, 5, 6, 7}
print(a | b)  # {1, 2, 3, 4, 5, 6, 7}
print(a & b)  # {3, 4}
print(a - b)  # {1, 2}
print(a ^ b)  # {1, 2, 5, 6, 7}
S = {-1, -5, 4} ALL ANY
print(any([x > 0 for x in S]))  # True
print(all([x > 0 for x in S]))  # False
x = 1, 2, 3, 4, 5, 6  UNPACKING
_, _, *y, _, _ = x
s = y
print(s)  # [3, 4]


class Node:
  LINKED LISTS

  def __init__(self, informatie=None):
    self.informatie = informatie
    self.adresa_urmatoare = None


class LinkedList:
  def __init__(self):
    self.informatie_inceput = None

  def printing(self):
    printvalue = self.informatie_inceput
    while printvalue is not None:
      print(printvalue.informatie)
      printvalue = printvalue.adresa_urmatoare

  def ins_beg(self, o_informatie):
    new = Node(o_informatie)
    new.adresa_urmatoare = self.informatie_inceput
    self.informatie_inceput = new

  def ins_end(self, o_informatie):
    newest = Node(o_informatie)
    if self.informatie_inceput == None:
      self.informatie_inceput = newest
      return
    lastNode = self.informatie_inceput
    while(lastNode.adresa_urmatoare):
      lastNode = lastNode.adresa_urmatoare
    lastNode.adresa_urmatoare = newest

  def ins_bet(self, node, o_informatie):
    newofall = Node(o_informatie)
    newofall.adresa_urmatoare = node.adresa_urmatoare
    node.adresa_urmatoare = newofall

  def deletion(self, informatia_stearsa):
    new = self.informatie_inceput
    if new is not None:
      if new.informatie == informatia_stearsa:
        self.informatie_inceput = new.adresa_urmatoare
        new = None
        return

    while new is not None:
      if new.informatie == informatia_stearsa:
        break
      prev = new
      new = new.adresa_urmatoare

    if new == None:
      return

    prev.adresa_urmatoare = new.adresa_urmatoare
    new = None


x = LinkedList()
x.informatie_inceput = Node('Manchester City')
data2 = Node('Liverpool')
data3 = Node('Tottenham')

x.informatie_inceput.adresa_urmatoare = data2
data2.adresa_urmatoare = data3
x.ins_beg('Manchester United')
x.ins_end('Everton')
x.ins_bet(x.informatie_inceput.adresa_urmatoare.adresa_urmatoare, 'Chelsea')
x.ins_bet(x.informatie_inceput.adresa_urmatoare.adresa_urmatoare, 'Barcelona')
x.deletion('Barcelona')
x.printing()


class Operations():
  MAGIC METHOD

  def __init__(self, *args):
    if len(args) == 0:
      self.numbers = (0, 0)
    else:
      self.numbers = args

  def __add__(self, other):
    sum = tuple(x + y for x, y in zip(self.numbers, other.numbers))
    return Operations(*sum)

  def __mul__(self, other):
    mul = tuple(x * y for x, y in zip(self.numbers, other.numbers))
    return Operations(*mul)


# also add/sub/mul/div/floor/mod/power/lshift/rshift/and/or/xor
x = Operations(4, 5)
y = Operations(1, 4)
t = Operations(3, 2)
z = x + y + t
w = x * y * t
print(z.numbers)
print(w.numbers)


class Comparison():
  MAGIC METHOD

  def __init__(self, x):
    self.x = x

  def __lt__(self, other):
    return self.x < other.x

  def __gt__(self, other):
    return self.x > other.x

  def __eq__(self, other):
    return self.x == other.x


if __name__ == '__main__':
  obj1 = Comparison(2)
  obj2 = Comparison(3)
  print(obj1 < obj2)  # True
  print(obj1 > obj2)  # False
  print(obj1 == obj2)  # False
# also iadd/isub/imul/idiv/ifloordivi/mod/ipower/ilshift/irshift/iand/ior/ixor  MAGIC METHOD


class dictionary(dict):
  MAGIC METHOD

  def __add__(self, other):
    self.update(other)
    return dictionary(self)


dict1 = dictionary({'firstname': 'Dragos'})
dict2 = dictionary({'lastname': 'BRATU'})
print(dict1 + dict2)


class LenthConversion:
  MAGIC METHOD

  value = {'mm': 0.001, 'cm': 0.01, 'm': 1, 'km': 1000, 'in': 0.254, 'ft': 0.3048,
           'yd': 0.9144, 'mi': 1609.344}

  def __init__(self, x, value_unit='m'):
    self.x = x
    self.value_unit = value_unit

  def Convert_to_meters(self):
    return self.x * LenthConversion.value[self.value_unit]

  def __add__(self, other):
    ans = self.Convert_to_meters() + other.Convert_to_meters()
    return LenthConversion(ans / LenthConversion.value[self.value_unit], self.value_unit)

  def __str__(self):
    return str(self.Convert_to_meters)

  def __repr__(self):
    return "LengthConversion(" + str(self.x) + ' ' + self.value_unit + ")"


obj1 = LenthConversion(0, 'm') + LenthConversion(52, 'cm')
print(repr(obj1))
print(obj1)


def teams(*args, **kwargs): *args
  print(args)
  print(kwargs)


england = ['United', 'Chelsea']
spain = {'spain_1': 'Barcelona', 'spain_2': 'Real'}
teams('United', "Chelsea", spain_1='Barcelona', spain_2='Real')
teams(*england, **spain)
pip install virtualenv  virtualenv
virtualenv project1_env
source project1_env / bin / activate
pip freeze - -local > requirements.txt
deactivate
pip install - r requirements.txt
nums = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9] generator creator
my_gen = (n * n for n in nums)


class Angajat():
  sort objects

  def __init__(self, nume, varsta, salariu):
    self.nume = nume
    self.varsta = varsta
    self.salariu = salariu

  def __repr__(self):
    return f'({self.nume}, {self.varsta}, {self.salariu})'


a1 = Angajat('Dragos', 25, 60000)
a2 = Angajat('Vlad', 28, 80000)
a3 = Angajat('Iulian', 24, 30000)

angajati = [a1, a2, a3]

ang_sort_sal = sorted(angajati, key=lambda e: e.salariu, reverse=True)
print(ang_sort_sal)
import os os module

print(os.getcwd())
os.chdir('D:\\LocalData\\py06366\\OneDrive - Alliance')
print(os.removedirs('D:\\LocalData\\py06366\\OneDrive - Alliance\\MaaanU'))
import os os.stat
from datetime import datetime

mod_time = os.stat('Server.py').st_mtime
print(datetime.fromtimestamp(mod_time))
import os os.walk

for dir_path, dir_names, file_names in os.walk('C:\\Users\\py06366\\OneDrive - Alliance\\Pyhon\\a\\PyCharm'):
  print('Current Path:', dir_path)
  print('Directories:', dir_names)
  print('Files:', file_names)
with open('client1.py', 'r+') as f:
  OPEN BIG FILES
  size_to_read = 10
  f_contents = f.read(size_to_read)
  while len(f_contents) > 0:
    print(f_contents, end='#')
    f_contents = f.read(size_to_read)
with open('client1.py', 'r+') as f:
  SEEK
  size_to_read = 10
  f_contents = f.read(size_to_read)
  print(f_contents)
  f.seek(0)
  f_contents = f.read(size_to_read)
  print(f_contents)
import random RANDOM.UNIFORM

value = random.uniform(1, 10)
print(value)  # 4.15449409933993
import random RANDOM.CHOICES

colors = ['Red', 'Black', 'Green']
roulette = random.choices(colors, weights=[18, 18, 2], k=5)
print(roulette)  # ['Red', 'Black', 'Black', 'Red', 'Black']
import os RENAME FILES

os.chdir('C:\\Users\\py06366\\OneDrive - Alliance\\Pyhon\\P1\\moby\\Fisiere')
for f in os.listdir():
  file_name, file_extention = os.path.splitext(f)
  echipa, categoria, numar = file_name.split('-')
  echipa = echipa.strip()
  categoria = categoria.strip()
  numar = numar.strip()
  nume_nou = f'{numar}-{echipa}{file_extention}'
  os.rename(f, nume_nou)
. - Any character except new line REGEX
\d - digit(0 - 9)
\D - Not a digit(0 - 9)
\w - Word character(a - z, A - Z, 0 - 9, _)
\W - Not a word character(a - z, A - Z, 0 - 9, _)
\s - Whitespaces(space, tab, newline)
\S - Not whitespaces(space, tab, newline)
\b - Word boundary
\B - Not word boundary
^ - Beginning of a string
$ - End of a string

Quantifiers:
* - 0 or More
+ - 1 or More
? - 0 or One
{3} - Exact number
{3, 4} - Range of numbers(min, max)

Groups:
| - Either or
() - Group
for n in range(1, 11):
  ZEROS
  print(f'the value is {n:03}')
from functools import wraps DECORATORS


def my_logger(original_function):
  import logging
  logging.basicConfig(filename=f'{original_function.__name__}.log', level=logging.INFO)

  @wraps(original_function)
  def wrapper(*args, **kwargs):
    logging.info(f'Ran with args: {args}, and {kwargs}')
    return original_function(*args, **kwargs)
  return wrapper


def my_timer(original_function):
  import time

  @wraps(original_function)
  def wrapper(*args, **kwargs):
    t1 = time.time()
    result = original_function(*args, **kwargs)
    t2 = time.time() - t1
    print(f'{original_function.__name__} ran in : {t2} seconds.')
    return result
  return wrapper


import time


@my_logger
@my_timer
def display(echipa, trofee):
  time.sleep(1)
  print(f'Display ran with arguments {echipa}, {trofee}')


display('Manchester', 20)
from datetime import datetime DATETIME

bday = datetime(1994, 4, 19)
print(f'My bday is on {bday:%B %d, %Y}')  # My bday is on April 19, 1994
pi = 3.14159265 DELIMITER
print(f'pi este {pi:.3f}')
import sqlite3  SQLITE3

# Class to create employees


class Employee:

  def __init__(self, first, last, pay):
    self.first = first
    self.last = last
    self.pay = pay

  @property
  def email(self):
    return f'{self.first}.{self.last}@email.com'

  @property
  def fullname(self):
    return f'{self.first} {self.last}'

  def __repr__(self):
    return f"Employee ('{self.first}','{self.last}', {self.pay})"


# Connection to a datebase (file or memory)
conn = sqlite3.connect(':memory:')
c = conn.cursor()
# Creation of a table "employees"
c.execute("""CREATE TABLE employees ( 
            first text, 
            last text,  
            pay integer 
            )""")

# creation of 3 functions: insertion, get employee by name, update salary and remove employee


def insert_emp(emp):
  with conn:
    c.execute(f"INSERT INTO employees VALUES (?, ?, ?)", (emp.first, emp.last, emp.pay))


def get_emp_by_name(lastname):
  c.execute("SELECT * FROM employees WHERE last=:last", {"last": lastname})
  return c.fetchall()


def update_pay(emp, pay):
  with conn:
    c.execute("""UPDATE employees SET pay = :pay  
                    WHERE first =:first AND last = :last""",
              {'first': emp.first, 'last': emp.last, 'pay': pay})


def remove_emp(emp):
  with conn:
    c.execute("DELETE FROM employees WHERE first=:first AND last=:last", {
              "first": emp.first, 'last': emp.last})


# Creation of 2 employee objects
emp_1 = Employee('Marian', "POPA", 45000)
emp_2 = Employee('Mara', "POPA", 35000)

# Insertion into the database
insert_emp(emp_1)
insert_emp(emp_2)
# printing
emps = get_emp_by_name('POPA')
print(emps)

# update one salary and remove one employee
update_pay(emp_2, 99999)
remove_emp(emp_1)
emps = get_emp_by_name('POPA')
print(emps)
conn.close()
import pandas as pd IMPORTING PANDAS
df = pd.read_csv('pokemon_data.csv', delimiter=',', chunksize=5)  READ  A CSV FILE
print(df.tail(10))  # Name   Type 1  ... Speed  Generation  Legendary 797  720  HoopaHoopa Confined  Psychic  ...    70           6       True 798  720   HoopaHoopa Unbound  Psychic  ...    80           6       True 799  721            Volcanion     Fire  ...    70           6       True  READ THE TAIL/HEAD
# to print the header Index(['#', 'Name', 'Type 1', 'Type 2', 'HP', 'Attack', 'Defense', 'Sp. Atk',        'Sp. Def', 'Speed', 'Generation', 'Legendary'],       dtype='object')  READ THE HEADER
print(df.columns)
print(df['Name'][0:5])  # to print all the value of a column 0                Bulbasaur 1                  Ivysaur 2                 Venusaur 3    VenusaurMega Venusaur 4               Charmander Name: Name, dtype: object READ A COLUMN
print(df[['Name', 'Type 1', 'HP']])  # to print all the value of a column                       Name   Type 1  HP 0                Bulbasaur    Grass  45 1                  Ivysaur    Grass  60 2                 Venusaur    Grass  80 3    VenusaurMega Venusaur    Grass  80 4               Charmander     Fire  39 ..                     ...      ...  .. 795                Diancie     Rock  50 796    DiancieMega Diancie     Rock  50 797    HoopaHoopa Confined  Psychic  80 798     HoopaHoopa Unbound  Psychic  80 799              Volcanion     Fire  80 READ MULTIPLE COLUMNS
print(df.iloc[1:4])  # Name Type 1  Type 2  ...  Sp. Def  Speed  Generation  Legendary 1  2   Ivysaur  Grass  Poison  ...       80     60           1      False 2  3  Venusaur  Grass  Poison  ...      100     80           1      False  PRINT EACH ROW
print(df.iloc[2, 1]) Venusaur PRINT A SPECIFIC POSITION
for index, row in df.iterrows():
  ITERATE THROUGH EACH ROW
  print(index, row['Name'], row['HP'], row['Attack']) 0 Bulbasaur 45 49 1 Ivysaur 60 62 2 Venusaur 80 82 3 VenusaurMega Venusaur 80 100
# Name Type 1  ... Speed  Generation  Legendary 4      4                 Charmander   Fire  ...    65           1      False 5      5                 Charmeleon   Fire  ...    80           1      False PRINT SPECIFIC CATEGORY
print(df.loc[df['Type 1'] == 'Fire'])
print(df.describe())  # HP      Attack  ...     Sp. Def       Speed  Generation count  800.000000  800.000000  800.000000  ...  800.000000  800.000000   800.00000 mean   362.813750   69.258750   79.001250  ...   71.902500   68.277500     3.32375 std    208.343798   25.534669   32.457366  ...   27.828916   29.060474     1.66129 min      1.000000    1.000000    5.000000  ...   20.000000    5.000000     1.00000 25%    184.750000   50.000000   55.000000  ...   50.000000   45.000000     2.00000 50%    364.500000   65.000000   75.000000  ...   70.000000   65.000000     3.00000 75%    539.250000   80.000000  100.000000  ...   90.000000   90.000000     5.00000 max    721.000000  255.000000  190.000000  ...  230.000000  180.000000     6.00000 DETAILS
# Name Type 1  ... Speed  Generation  Legendary 726  658              Greninja  Water  ...   122           6      False 130  121               Starmie  Water  ...   115           1      False 466  419              Floatzel  Water  ...   115           4      False SORT
print(df.sort_values(['Type 1', 'Speed'], ascending=(0, 0)))
df['Total'] = df['HP'] + df['Attack'] + df['Defense'] + df['Sp. Atk'] + \ NEW COLUMN
  df['Sp. Def'] + df['Speed']
df['Total'] = df.iloc[:, 4:10].sum(axis=1)  SAU NEW COLUMN
df = df.drop(columns=['Total']) DROP A COLUMN
df['Total'] = df.iloc[:, 4:10].sum(axis=1)  ADD A COLUMN TO A SPECIFIC PLACE
columns = list(df.columns.values)
df = df[columns[0:4] + [columns[-1]] + columns[4:12]]
print(df.head(5))  # Name Type 1  ... Speed  Generation  Legendary 0  1              Bulbasaur  Grass  ...    45           1      False 1  2                Ivysaur  Grass  ...    60           1      False 2  3               Venusaur  Grass  ...    80           1      False
df.to_csv('modified.csv', index=False)  # csv SAVE THE NEW CSV
df.to_csv('modified.txt', index=False, sep='\t')  # txt
# Name Type 1  ... Speed  Generation  Legendary 2      3               Venusaur  Grass  ...    80           1      False 3      3  VenusaurMega Venusaur  Grass  ...    80           1      False 50    45              Vileplume  Grass  ...    50           1      False 77    71             Victreebel  Grass  ...    70           1      False 652  591              Amoonguss  Grass  ...    30           5      False  PRINT WITH ADVANCED FILTERING
print(df.loc[(df['Type 1'] == 'Grass') & (
    df['Type 2'] == 'Poison') & (df['HP'] > 70)])
new_df = new_df.reset_index(drop=True)  RESET INDEX
new_df.reset_index(drop=True, inplace=True) RESET INDEX INPLACE
print(df.loc[(df['Name'].str.contains('Mega'))])  # Name    Type 1  ... Speed  Generation  Legendary 3      3      VenusaurMega Venusaur     Grass  ...    80           1      False 7      6  CharizardMega Charizard X      Fire  ...   100           1      False 8      6  CharizardMega Charizard Y      Fire  ...   100           1      False  CONTAINS FILTER
# Name   Type 1  ... Speed  Generation  Legendary 0      1            Bulbasaur    Grass  ...    45           1      False 1      2              Ivysaur    Grass  ...    60           1      False 2      3             Venusaur    Grass  ...    80           1      False  ~ FOR NOT CONTAINING
print(df.loc[~df['Name'].str.contains('Mega')])
# Name Type 1  ... Speed  Generation  Legendary 0      1              Bulbasaur  Grass  ...    45           1      False 1      2                Ivysaur  Grass  ...    60           1      False 2      3               Venusaur  Grass  ...    80           1      False  USING REGEX
print(df.loc[df['Type 1'].str.contains('fire|grass', flags=re.I, regex=True)])
re.I  # Ignore case IGNORE CASE
df.loc[df['Type 1'] == 'Fire', 'Type 1'] = 'Flamer'  # Name   Type 1  ... Speed  Generation  Legendary 0      1              Bulbasaur    Grass  ...    45           1      False 1      2                Ivysaur    Grass  ...    60           1      False 2      3               Venusaur    Grass  ...    80           1      False 3      3  VenusaurMega Venusaur    Grass  ...    80           1      False 4      4             Charmander   Flamer  ...    65           1      False CHANGE VALUES
# Name   Type 1  ... Speed  Generation  Legendary 0      1              Bulbasaur    Grass  ...    45           1      False 1      2                Ivysaur    Grass  ...    60           1      False 2      3               Venusaur    Grass  ...    80  TEST VALUE       EHHE 3      3  VenusaurMega Venusaur    Grass  ...    80  TEST VALUE       EHHE MULTIPLE CHANGE
df.loc[df['Total'] > 500, ['Generation', 'Legendary']] = ['TEST VALUE', 'EHHE']
# .mean/.sum/.count                    #       Total         HP  ...       Speed  Generation  Legendary Type 1                                       ...                                    Steel     442.851852  487.703704  65.222222  ...   55.259259    3.851852   0.148148 Rock      392.727273  453.750000  65.363636  ...   55.909091    3.454545   0.090909 Dragon    474.375000  550.531250  83.312500  ...   83.031250    3.875000   0.375000 GROUP BY
print(df.groupby(['Type 1']).mean().sort_values('Defense', ascending=False))
df['Count'] = 1 NEW COLUMN FOR COUNT
print(df.groupby(['Type 1', 'Type 2']).count()['Count']) Type 1  Type 2   Bug     Electric     2         Fighting     2         Fire         2         Flying      14         Ghost        1                     .. Water   Ice          3         Poison       3         Psychic      5         Rock         4         Steel        1
for df in pd.read_csv('pokemon_data.csv', chunksize=5):
  CHUNKING THE FILE
import numpy as np  IMPORTING NUMPY

a = np.array([(1, 2, 3), (4, 5, 6)])
print(a)
s = range(1000) LESS MEMORY
print(sys.getsizeof(5) * len(s))  # 14000
d = np.arange(1000)
print(d.size * d.itemsize)  # 4000
import numpy as np  FASTER
import time

l1 = range(1000000)
l2 = range(1000000)
a1 = np.arange(1000000)
a2 = np.arange(1000000)

start = time.time()
result = [(x, y) for x, y in zip(l1, l2)]
print((time.time() - start) * 1000)  # 310.00423431396484

start = time.time()
result = a1 + a2
print((time.time() - start) * 1000)  # 65.00744819641113
a = np.array([(1, 2, 3), ('a', 32, 3)])
print(a.ndim) NDIM
print(a.itemsize) ITEMSIZE
print(a.dtype)  DTYPE
print(a.size) SIZE
print(a.shape)  SHAPE
print(a.reshape(6, 1))  RESHAPE
a = np.array([(1, 2, 3, 4, 5),  SLICING
              ('a', 'b', 'c', 'd', 'e'),
              (11, 22, 33, 44, 55),
              ('A', ' B', 'C', 'D', 'E'),
              (6, 7, 8, 9, 0),
              ('aa', 'bb', 'cc', 'dd', 'ee')])
print(a[1:3, 2:4])[['c' 'd']['33' '44']]
a = np.linspace(0, 10, 6) LINSPACE
print(a)  # [ 0.  2.  4.  6.  8. 10.]
print(a.max())  MAX
print(a.sum())  SUM
a = np.array([(1, 2, 3, 4, 5),  AXIS SUM
              (11, 22, 33, 44, 55),
              (6, 7, 8, 9, 0)])

print(a.sum(axis=0))[18 31 44 57 60]
print(a.sum(axis=1))[15 165  30]
print(np.sqrt(a))[[1.         1.41421356 1.73205081 2.         2.23606798][3.31662479 4.69041576 5.74456265 6.63324958 7.41619849][2.44948974 2.64575131 2.82842712 3.         0.]] SQUARE ROOT
print(np.std(a)) 16.32993161855452  STANDARD DEVIATION
a = np.array([(4, 4, 4),  OPERATIONS
              (3, 2, 1)])
b = np.array([(2, 2, 3),
              (3, 2, 1)])

print(a - b)[[2 2 1][0 0 0]]
print(np.vstack((a, b)))  VERTICAL / HORIZONTAL STACKING / CONCATENATION
print(np.hstack((a, b)))
a = np.array([(4, 4, 4),  ALL ELEMENTS IN A ROW
              (3, 2, 1)])
print(a.ravel())[4 4 4 3 2 1]
x = np.arange(0, 3 * np.pi, 0.1)  SINUS COSINUS
y = np.sin(x)
y = np.cos(x)
ar = np.array([1, 2, 3])  EXP LOG
print(np.exp(ar))[2.71828183  7.3890561  20.08553692]
print(np.log10(ar))[0.         0.30103    0.47712125]
arr = np.array([1, 2, 3, 4, 5, 6, 7, 8, 9]) FILTERING ITEMS
print(arr[arr % 2 == 1])  # [1 3 5 7 9]
arr = np.array([1, 2, 3, 4, 5, 6, 7, 8, 9]) FILTER AND REPLACE
arr[arr % 2 == 1] = -1
print(arr)  # [-1  2 -1  4 -1  6 -1  8 -1]
a = np.arange(10) FILTER AND REPLACE
b = np.where(a % 2 == 0, -1, a)
print(b)  # [-1  1 -1  3 -1  5 -1  7 -1  9]
arr = np.array([1, 2, 3, 4, 5, 6, 7, 8, 9, 10]) RESHAPE
b = np.reshape(arr, (2, -1))
print(b)
# [[ 1  2  3  4  5]
#  [ 6  7  8  9 10]]
a = np.arange(10).reshape(2, -1)  STACKING
b = np.repeat(1, 10).reshape(2, -1)
c = np.vstack([a, b])  # hstack
c = np.concatenate([a, b], axis=1)  # axis=0
c = np.r_[a, b]  # c_[a, b]
print(c)
a = np.array([1, 2, 3]) TILE
b = np.r_[np.repeat(a, 3), np.tile(a, 3)]
print(b)  # [1 1 1 2 2 2 3 3 3 1 2 3 1 2 3 1 2 3]
a = np.array([1, 2, 3, 2, 3, 4, 3, 4, 5, 6])  COMMON ELEMENTS
b = np.array([7, 2, 10, 2, 7, 4, 9, 4, 9, 8])
print(np.intersect1d(a, b))  # [2 4]
a = np.array([1, 2, 3, 3])  DIFFERENT ELEMENTS
b = np.array([3, 4, 5])
print(np.setdiff1d(a, b))  # [1 2]
a = np.array([1, 2, 3, 3, 6, 7])  COINCIDENCE INDEX
b = np.array([3, 4, 5, 3, 4, 7])
print(np.where(a == b))  # (array([3, 5]),)
a = np.array([2, 6, 1, 9, 10, 3, 27]) FILTER
index = np.where((a >= 5) & (a <= 10))
print(a[index])  # [ 6  9 10]
print(a[(a >= 5) & (a <= 10)])  # [ 6  9 10]
a = np.array([5, 7, 9, 8, 6, 4, 5]) FUNCTION APLICATION
b = np.array([6, 3, 4, 8, 9, 7, 1])


def ceva(x, y):
  return x * 10 + y


pair_max = np.vectorize(ceva)
print(pair_max(a, b))  # [56 73 94 88 69 47 51]
a = np.arange(9).reshape((3, 3))  SWAP COLUMNS / ROWS
print(a[:, [1, 0, 2]])
print(a[[1, 0, 2], :])
a = np.arange(9).reshape((3, 3))  REVERSE ROWS
print(a[::-1])
a = np.arange(9).reshape((3, 3))  REVERSE COLUMNS
print(a[:, ::-1])
